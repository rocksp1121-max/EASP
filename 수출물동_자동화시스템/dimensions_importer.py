# -*- coding: utf-8 -*-
"""
dimensions_importer.py — 모델 치수(LWH) 마스터 import

입력: SAP export Excel (73컬럼, 시트 'Data')
출력: data/ref_dimensions.csv (Model 단위 unique)
규칙:
  - Model 신규 → INSERT
  - Model 기존 + 기존 L_mm=0 → UPDATE (dim 보강)
  - Model 기존 + 기존 L_mm>0 → skip (사용자: "바뀌는거보다 추가되는 구조")
"""
import csv
import datetime
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).parent / "data"
REF_PATH = DATA_DIR / "ref_dimensions.csv"
LOG_PATH = DATA_DIR / "ref_dimensions_log.csv"

FIELDS = ["Model", "L_mm", "W_mm", "H_mm", "Volume_m3", "Weight_kg", "Dept", "UpdatedAt", "Source"]


def _load_existing():
    if not REF_PATH.exists():
        return {}
    with open(REF_PATH, encoding="utf-8-sig", newline="") as f:
        return {r["Model"]: r for r in csv.DictReader(f) if r.get("Model")}


def _save(records):
    REF_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(REF_PATH, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        for m in sorted(records.keys()):
            w.writerow(records[m])


def _append_log(source, stats):
    is_new = not LOG_PATH.exists()
    with open(LOG_PATH, "a", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        if is_new:
            w.writerow(["When", "Source", "Inserted", "Updated", "Skipped", "TotalRows"])
        w.writerow([datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    source, stats["inserted"], stats["updated"], stats["skipped"], stats["total"]])


def _to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def import_file(file_path, existing=None):
    """
    1개 Excel 파일 import. existing dict 주면 그걸 직접 수정, 아니면 csv 로드.
    반환: {inserted, updated, skipped, total, source}
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]   # 'Data'

    # 헤더 매핑
    hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx = {h: i for i, h in enumerate(hdr_row) if h}

    required = ("Model", "Gross Width", "Gross Height", "Gross Length")
    for k in required:
        if k not in col_idx:
            raise ValueError(f"필수 컬럼 누락: {k}")

    save_now = existing is None
    if save_now:
        existing = _load_existing()

    source = Path(file_path).stem
    today = datetime.date.today().isoformat()

    stats = {"inserted": 0, "updated": 0, "skipped": 0, "total": 0, "source": source}

    vol_i  = col_idx.get("Gross Volume", -1)
    wt_i   = col_idx.get("Gross Weight", -1)
    dept_i = col_idx.get("Prod.H1.T", -1)
    m_i    = col_idx["Model"]
    l_i    = col_idx["Gross Length"]
    w_i    = col_idx["Gross Width"]
    h_i    = col_idx["Gross Height"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["total"] += 1
        model = (row[m_i] or "")
        if not isinstance(model, str):
            model = str(model)
        model = model.strip()
        if not model:
            continue

        L = _to_int(row[l_i])
        W = _to_int(row[w_i])
        H = _to_int(row[h_i])
        if not (L > 0 and W > 0 and H > 0):
            continue   # dim 없는 행은 csv 저장 안 함

        vol  = _to_float(row[vol_i])  if vol_i  >= 0 else 0.0
        wt   = _to_float(row[wt_i])   if wt_i   >= 0 else 0.0
        dept = (row[dept_i] or "")    if dept_i >= 0 else ""
        if not isinstance(dept, str):
            dept = str(dept)

        rec = {
            "Model": model,
            "L_mm": L, "W_mm": W, "H_mm": H,
            "Volume_m3": round(vol, 6),
            "Weight_kg": round(wt, 3),
            "Dept": dept.strip(),
            "UpdatedAt": today,
            "Source": source,
        }

        if model in existing:
            old = existing[model]
            try:
                old_L = int(float(old.get("L_mm", 0) or 0))
            except (TypeError, ValueError):
                old_L = 0
            if old_L > 0:
                stats["skipped"] += 1
                continue
            existing[model] = rec
            stats["updated"] += 1
        else:
            existing[model] = rec
            stats["inserted"] += 1

    wb.close()

    if save_now:
        _save(existing)
        _append_log(source, stats)

    return stats


def import_files(file_paths):
    """여러 파일 누적 import (한 번만 save). 합산 + 파일별 통계 반환."""
    existing = _load_existing()
    agg = {"inserted": 0, "updated": 0, "skipped": 0, "total": 0, "files": []}
    for fp in file_paths:
        s = import_file(fp, existing=existing)
        agg["inserted"] += s["inserted"]
        agg["updated"]  += s["updated"]
        agg["skipped"]  += s["skipped"]
        agg["total"]    += s["total"]
        agg["files"].append({
            "source":   s["source"],
            "inserted": s["inserted"],
            "updated":  s["updated"],
            "skipped":  s["skipped"],
            "total":    s["total"],
        })
        _append_log(s["source"], s)
    _save(existing)
    return agg


def stats():
    """현재 ref_dimensions.csv 통계."""
    if not REF_PATH.exists():
        return {"total": 0, "by_dept": {}}
    by_dept = {}
    total = 0
    with open(REF_PATH, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            total += 1
            d = r.get("Dept") or "(미분류)"
            by_dept[d] = by_dept.get(d, 0) + 1
    return {"total": total, "by_dept": by_dept}


def lookup(model):
    """모델 단건 조회 (LWH 등 dict 반환, 없으면 None). 메모리 캐시 사용."""
    cache = _lookup_cache()
    return cache.get((model or "").strip())


_CACHE = {"data": None, "mtime": 0}

def _lookup_cache():
    if not REF_PATH.exists():
        return {}
    mt = REF_PATH.stat().st_mtime
    if _CACHE["data"] is not None and _CACHE["mtime"] == mt:
        return _CACHE["data"]
    d = {}
    with open(REF_PATH, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            m = (r.get("Model") or "").strip()
            if m:
                d[m] = r
    _CACHE["data"] = d
    _CACHE["mtime"] = mt
    return d


def reload():
    _CACHE["data"] = None
    _CACHE["mtime"] = 0
